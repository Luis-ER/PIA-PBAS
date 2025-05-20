import numpy, statistics, collections, re
from openpyxl import Workbook
import matplotlib.pyplot as plt

i = 0
razas = []
contador = []
rep = []
regex1 = re.compile(r"[a-zA-Z]+\-*[a-zA-Z]*")
regex2 = re.compile(r".+\.jpg")
longitud=[]
with open ("Caninos.txt", "r") as archivaldo:
    for linea in archivaldo:
        conf = False
        if i%2 == 0:
            print("Raza")
            if re.match(regex1, linea):
                razas.append(linea.rstrip("\n"))
                conf = True
        else:
            print("Foto")
            if re.match(regex2, linea):
                conf = True
        if conf == True:
            print(linea)
        else:
            print("Error en los datos.")
        i += 1

razas_r = razas.copy()
razas_r.reverse()

for i in range(len(razas)):
    check = rep.count(i)
    if check == 0:
        cont = 0
        for j in range(len(razas)):
            if razas[i] == razas[j]:
                cont += 1
                rep.append(j)
        if cont > 1:
            for j in range(cont-1):
                razas_r.remove(razas[i])
        contador.append(cont)
            
razas_r.reverse()

print("Razas",razas_r)
print("Contador",contador)

razas_longitud=razas.copy()
for i in razas_longitud:
    long=len(i)
    longitud.append(long)

repeticiones={}
for long in longitud:
    if long in repeticiones:
        repeticiones[long]+=1
    else:
        repeticiones[long]=1
print("Repeticiones por Longitud",repeticiones)

totalint=len(razas)
contrepeticion=[]
intentosrep=0
for i in razas:
    intentosrep+=1
    if i in contrepeticion:
        print("Intentos Totales para Llegar a la Primera Repetición",intentosrep)
        rep = True
        break
    else:
        contrepeticion.append(i)
if rep != True:
    intentosrep = 0

contador2 = contador.copy()
contador2.sort
if contador2[0] != contador2[1]:
    u_max = True
    maximo_index = contador.index(contador2[0])
    maximo_raza = razas_r[maximo_index]
else:
    maximo_raza = "Más de una raza comparte el máximo"

x = [maximo_raza, "Intentos Totales"]
y = [contador2[0], len(razas)]
plt.bar(x,y)
plt.title("Raza más repetida vs Cantidad de intentos")
plt.ylabel("Cantidad")
plt.savefig("RazaRepvsCantidad.png")
plt.show()
    
x=numpy.array(razas_r)
y=numpy.array(contador)
plt.pie(y,labels=x)
plt.legend(x)
plt.title('Razas vs Cantidad')
plt.savefig("RazasvsCantidad_grafpie.png")
plt.show()

plt.scatter(x,y)
plt.colorbar()
plt.title("Razas vs Cantidad")
plt.xlabel("Razas")
plt.ylabel("Cantidad")
plt.savefig("RazasvsCantidad_grafdispersa")
plt.show()

x=repeticiones.keys()
y=repeticiones.values()
plt.bar(x,y)
plt.title('Repeticiones de Longitud de Nombres')
plt.xlabel('Cantidad de Letras')
plt.ylabel('Repeticiones')
plt.savefig("RepeticioneLongitud_grafbarra.png")
plt.show()


x = ["Primera Rep", "Intentos Totales"]
y = [intentosrep, totalint]
plt.bar(x,y)
plt.title('Intentos Totales para Llegar a la Primera Repetición')
plt.xlabel('Intentos')
plt.ylabel('Repetición Lograda')
plt.savefig("IntentosPrimeraRepetición_graflinea.png")
plt.show()

media=statistics.mean(contador)
mediana=statistics.median(contador)
moda=statistics.mode(contador)
desviacionestandar=statistics.pstdev(contador)
varianza=statistics.pvariance(contador)
distribucion=statistics.stdev(contador)

print("Media:",media)
print("Mediana:",mediana)
print("Moda:",moda)
print("Desviacion Estandar:",desviacionestandar)
print("Varianza:",varianza)
print("Distribución:",distribucion)

#pitbull, terrier-toy, doberman, mastiff-english, shihtzu
#bakharwal-indian, schipperke, hound-blood, chow, maltese

libro=Workbook()
hoja=libro.active
cadena=("Caninos")
hoja.title=cadena
hoja["B1"]="Razas"
hoja["C1"]="Contador"
print (hoja["B1"].value,end = "\t")
print (hoja["C1"].value, end = "\t")

razas2=[]
contador2=[]
for perros in razas:
    for perritos in razas2:
        if perritos==perros:
            break
    else:
        razas2.append(perros)
        contador2.append(razas.count(perros))
        
count=2
for i in range(len(razas2)):
    hoja.cell(count,1,i+1)
    hoja.cell(count,2,razas2[i])
    hoja.cell(count,3,contador2[i])
    count+=1
libro.save("ExcelCaninos.xlsx")
